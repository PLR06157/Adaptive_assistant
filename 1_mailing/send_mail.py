"""
Utility for sending personalized HTML emails with attachments via Microsoft 365.

The script reads recipient data from an XLSX spreadsheet, renders an HTML template using
row values, and delivers the messages through the Microsoft Graph API.

Template Variables:
    HTML templates use $variable syntax for placeholders (e.g., $first_name, $email).
    Available variables: $email, $first_name, $login_id, $access_code, $subject, $sender_email
    $login_id and $access_code are read from the "Login ID" and "Access Code" XLS columns
    and rendered as empty strings when those cells are blank.
    This syntax is safe with CSS curly braces {} in your HTML.

Configuration:
    The script automatically loads settings from a .env file in the mailing/ directory.
    All command-line arguments have default values that can be set via environment variables:

    Required (in .env):
        TENANT_ID           - Azure AD tenant ID
        CLIENT_ID           - Azure AD application (client) ID
        CLIENT_SECRET       - Azure AD client secret
        SENDER_EMAIL        - Email address to send from
        MAIL_SUBJECT        - Fallback subject line (used when the spreadsheet has no
                              'Subject' column or a row's Subject cell is empty)

    Optional (in .env):
        XLSX_PATH           - Path to recipient spreadsheet (default: mailing/recipients.xlsx)
        HTML_TEMPLATE_PATH  - Path to email template (default: mailing/email_template.html)
        ATTACHMENT_PATH     - Path to file attachment (optional, no default)
        MIN_WAIT_SECONDS    - Minimum wait between emails (default: 5)
        MAX_WAIT_SECONDS    - Maximum wait between emails (default: 15)
        MAX_RETRIES         - Maximum retry attempts on transient errors (default: 5)
        LOG_LEVEL           - Logging level (default: INFO)
        SAVE_TO_SENT_ITEMS  - Save to Sent Items folder (default: true)
        CC_EMAIL            - Comma-separated email addresses to CC on every outgoing message (optional)
        CONTINUE_ON_ERROR   - Continue with other recipients if one fails (default: false)

    Command-line arguments override .env values.

Usage:

Pre-processing (required when template contains local gallery images):
    python3 mailing/prepare_email.py --template mailing/sets/<folder>/template.html

    Run this once before sending whenever you add or replace photos in a gallery.
    It stamps explicit pixel dimensions on <img> tags so Outlook Windows renders
    them correctly. Safe to re-run; skips images that already have fixed dimensions.

Basic execution (uses .env defaults):
    python3 mailing/send_mail.py

With parameters (overrides .env):
    python3 mailing/send_mail.py \
        --xlsx mailing/recipients.xlsx \
        --template mailing/email_template.html \
        --attachment mailing/document.pdf \
        --mail-subject "My subject" \
        --min-wait 3 \
        --max-wait 10

Test without sending (dry-run):
    python3 mailing/send_mail.py --dry-run

Without saving to Sent Items:
    python3 mailing/send_mail.py --no-save-to-sent-items

Working example:

python3 mailing/send_mail.py \
    --xlsx mailing/recipients.xlsx \
    --template mailing/gbs_lions_event_email.html \
    --mail-subject "[LAST CALL] GBS Lions' Talks in Warsaw: AI - Is It Already a Mainstream Tool? 26.11.25" \
    --min-wait 0.5 \
    --max-wait 1 \
    --dry-run

"""

from __future__ import annotations

import argparse
import base64
import json
import logging
import mimetypes
import os
import random
import re
import string
import time
import uuid
from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import requests
from openpyxl import load_workbook
from premailer import transform
from dotenv import load_dotenv
from msal import ConfidentialClientApplication
import cssutils
cssutils.log.setLevel(logging.CRITICAL)

GRAPH_SCOPE = ["https://graph.microsoft.com/.default"]
GRAPH_ENDPOINT = "https://graph.microsoft.com/v1.0"


class ConfigurationError(RuntimeError):
    """Raised when required configuration is missing."""


def _read_env(name: str, *, required: bool = True, default: Optional[str] = None) -> str:
    value = os.getenv(name, default)
    if required and not value:
        raise ConfigurationError(
            f"Missing required configuration for {name}. "
            "Set it in your environment or .env file."
        )
    return value or ""


def _env_flag(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _load_html_template(path: Path) -> str:
    try:
        with open(path, 'r') as f:
            html = f.read()

        inline_html = transform(html)

        with open(path, 'w') as f:
            f.write(inline_html)
        
        return path.read_text(encoding="utf-8")
    except FileNotFoundError as exc:
        raise ConfigurationError(f"HTML template not found: {path}") from exc


def _guess_mime_type(path: Path) -> str:
    mime, _ = mimetypes.guess_type(path.name)
    return mime or "application/octet-stream"


def _build_attachment(path: Path) -> Dict[str, str]:
    if not path.exists():
        raise ConfigurationError(f"Attachment file does not exist: {path}")
    file_bytes = path.read_bytes()
    content_bytes = base64.b64encode(file_bytes).decode("ascii")
    return {
        "@odata.type": "#microsoft.graph.fileAttachment",
        "name": path.name,
        "contentType": _guess_mime_type(path),
        "contentBytes": content_bytes,
    }


def _prepare_inline_images(html: str, asset_root: Path) -> Tuple[str, List[Dict[str, str]]]:
    """
    Identify local <img> references and convert them to inline attachments that
    Microsoft Graph can embed via CID references.
    """
    matches = set(re.findall(r'<img[^>]+src=["\']([^"\']+)["\']', html, flags=re.IGNORECASE))
    if not matches:
        return html, []

    replacements: Dict[str, str] = {}
    attachments: List[Dict[str, str]] = []

    for src in matches:
        if src.startswith(("cid:", "http://", "https://", "data:")):
            continue
        image_path = Path(src)
        if not image_path.is_absolute():
            image_path = asset_root / src
        if not image_path.exists():
            logging.warning("Referenced inline image not found: %s", image_path)
            continue
        content_id = f"{Path(src).stem}-{uuid.uuid4().hex}@inline"
        file_bytes = image_path.read_bytes()
        attachments.append(
            {
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": image_path.name,
                "contentType": _guess_mime_type(image_path),
                "contentBytes": base64.b64encode(file_bytes).decode("ascii"),
                "contentId": content_id,
                "isInline": True,
            }
        )
        replacements[src] = content_id

    if not replacements:
        return html, attachments

    def _replace(match: re.Match) -> str:
        prefix, quote, value, _ = match.groups()
        content_id = replacements.get(value)
        if not content_id:
            return match.group(0)
        return f"{prefix}{quote}cid:{content_id}{quote}"

    updated_html = re.sub(
        r'(<img[^>]*src=)(["\'])([^"\']+)(\2)',
        _replace,
        html,
        flags=re.IGNORECASE,
    )
    return updated_html, attachments


@dataclass
class Recipient:
    email: str
    first_name: str
    subject: str
    context: Dict[str, str]


class GraphMailer:
    def __init__(self, tenant_id: str, client_id: str, client_secret: str, sender: str) -> None:
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        self._sender = sender
        self._client = ConfidentialClientApplication(
            client_id=client_id,
            client_credential=client_secret,
            authority=authority,
        )

    def _get_token(self) -> str:
        token = self._client.acquire_token_silent(GRAPH_SCOPE, account=None)
        if not token:
            token = self._client.acquire_token_for_client(scopes=GRAPH_SCOPE)
        if "access_token" not in token:
            raise RuntimeError(f"Unable to acquire access token: {json.dumps(token, indent=2)}")
        return token["access_token"]

    def _is_transient_error(self, status_code: int, response_text: str) -> bool:
        """Check if an error is transient and should be retried."""
        # 4xx errors (except 429) are client errors and should NOT be retried
        # These are permanent failures: bad email, invalid recipient, etc.
        if 400 <= status_code < 500 and status_code != 429:
            return False

        # Common transient HTTP status codes
        if status_code in (429, 502, 503, 504):
            return True

        # Check for specific transient error codes in response
        transient_error_codes = {
            "ErrorMailboxMoveInProgress",
            "ErrorServerBusy",
            "ErrorTimeoutExpired",
            "ErrorInternalServerError",
            "ErrorTooManyObjectsOpened",
        }

        try:
            error_data = json.loads(response_text)
            error_code = error_data.get("error", {}).get("code", "")
            return error_code in transient_error_codes
        except (json.JSONDecodeError, AttributeError):
            return False

    def _is_invalid_recipient_error(self, response_text: str) -> bool:
        """Check if the error is due to an invalid recipient email address."""
        try:
            error_data = json.loads(response_text)
            error_code = error_data.get("error", {}).get("code", "")
            return error_code in ("ErrorInvalidRecipients", "ErrorNonExistentMailbox")
        except (json.JSONDecodeError, AttributeError):
            return False

    def _get_retry_wait_time(self, status_code: int, attempt: int, retry_after: Optional[str] = None) -> int:
        """Calculate wait time for retry based on error type and attempt number."""
        # Use Retry-After header if provided
        if retry_after and retry_after.isdigit():
            return int(retry_after)

        # Different backoff strategies for different error types
        if status_code == 429:  # Throttling
            # Exponential backoff for throttling
            return min(30, (2 ** attempt) * 5)
        elif status_code == 503:  # Service unavailable (e.g., mailbox move)
            # Longer wait for mailbox operations (30s, 60s, 90s, 120s, 180s)
            return min(30, 30 + (attempt * 30))
        else:  # 502, 504, or other transient errors
            # Moderate exponential backoff
            return min(30, (2 ** attempt) * 10)

    def _send_with_retry(
        self,
        recipient: Recipient,
        rendered_html: str,
        payload: Dict,
        headers: Dict,
        max_retries: int = 5,
    ) -> None:
        """Send email with automatic retry on transient errors."""
        last_error_msg = ""

        for attempt in range(max_retries):
            try:
                response = requests.post(
                    f"{GRAPH_ENDPOINT}/users/{self._sender}/sendMail",
                    headers=headers,
                    json=payload,
                    timeout=30,
                )
            except requests.exceptions.ConnectionError as e:
                logging.warning("Network error for %s: %s. Retrying in 10 seconds...", recipient.email, e)
                time.sleep(10)
                continue

            # Success
            if response.status_code < 300:
                logging.info("Sent mail to %s", recipient.email)
                return

            # Check if this is a transient error worth retrying
            if self._is_transient_error(response.status_code, response.text):
                retry_after = response.headers.get("Retry-After")
                wait_time = self._get_retry_wait_time(response.status_code, attempt, retry_after)

                # Extract error details for logging
                error_detail = "Unknown error"
                try:
                    error_data = json.loads(response.text)
                    error_detail = error_data.get("error", {}).get("message", response.text)
                except (json.JSONDecodeError, AttributeError):
                    error_detail = response.text[:200]  # Limit error message length

                logging.warning(
                    "Transient error (%d) for %s: %s. Waiting %d seconds before retry %d/%d",
                    response.status_code,
                    recipient.email,
                    error_detail,
                    wait_time,
                    attempt + 1,
                    max_retries,
                )

                last_error_msg = f"{response.status_code} {response.text}"
                time.sleep(wait_time)

                # Refresh token before retry
                headers["Authorization"] = f"Bearer {self._get_token()}"
                continue

            # Non-transient error - fail immediately with a clear message
            error_detail = response.text
            try:
                error_data = json.loads(response.text)
                error_code = error_data.get("error", {}).get("code", "")
                error_message = error_data.get("error", {}).get("message", "")
                error_detail = f"{error_code}: {error_message}" if error_code else error_message
            except (json.JSONDecodeError, AttributeError):
                pass

            # Provide specific message for invalid recipient errors
            if self._is_invalid_recipient_error(response.text):
                raise RuntimeError(
                    f"Invalid recipient email '{recipient.email}': {error_detail}"
                )

            # Generic permanent error
            raise RuntimeError(
                f"Permanent error sending to {recipient.email} ({response.status_code}): {error_detail}"
            )

        # Max retries exceeded for transient error
        raise RuntimeError(
            f"Failed to send mail to {recipient.email} after {max_retries} retries. "
            f"Last error: {last_error_msg}"
        )

    def send(
        self,
        recipients: Iterable[Recipient],
        html_template: str,
        *,
        inline_attachments: Optional[List[Dict[str, str]]] = None,
        attachment: Optional[Dict[str, str]] = None,
        min_wait: float = 5.0,
        max_wait: float = 15.0,
        max_retries: int = 5,
        dry_run: bool = False,
        save_to_sent_items: bool = True,
        cc_email: Optional[str] = None,
        continue_on_error: bool = False,
    ) -> None:
        total = 0
        success_count = 0
        failed_recipients: List[Tuple[str, str]] = []  # (email, error_message)
        inline_attachments = inline_attachments or []
        cc_recipient_entry = (
            [{"emailAddress": {"address": addr.strip()}} for addr in cc_email.split(",") if addr.strip()]
            if cc_email else None
        )
        last_send_timestamp: Optional[float] = None
        for recipient in recipients:
            total += 1
            template = string.Template(html_template)
            rendered_html = template.safe_substitute(recipient.context)
            if dry_run:
                logging.info(
                    "[DRY-RUN] Would send: \n Subject: '%s' - Email: %s - Name: [%s]",
                    recipient.subject,
                    recipient.email,
                    recipient.first_name,
                )
                success_count += 1
                continue

            # Get fresh token for each email to prevent expiration during long mailings
            # MSAL will use cached token if still valid, or refresh automatically
            token = self._get_token()

            payload = {
                "message": {
                    "subject": recipient.subject,
                    "body": {"contentType": "HTML", "content": rendered_html},
                    "toRecipients": [{"emailAddress": {"address": recipient.email}}],
                },
                "saveToSentItems": save_to_sent_items,
            }
            if cc_recipient_entry:
                payload["message"]["ccRecipients"] = deepcopy(cc_recipient_entry)
            message_attachments: List[Dict[str, str]] = []
            if inline_attachments:
                message_attachments.extend(deepcopy(inline_attachments))
            if attachment:
                message_attachments.append(deepcopy(attachment))
            if message_attachments:
                payload["message"]["attachments"] = message_attachments
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            }

            # Send with automatic retry on transient errors
            try:
                self._send_with_retry(recipient, rendered_html, payload, headers, max_retries)
                success_count += 1
            except RuntimeError as exc:
                error_msg = str(exc)
                failed_recipients.append((recipient.email, error_msg))
                logging.error("Failed to send to %s: %s", recipient.email, error_msg)

                if not continue_on_error:
                    # Stop processing and re-raise the error
                    raise

                # Continue with next recipient
                logging.info("Continuing with remaining recipients...")

            if last_send_timestamp is not None:
                elapsed = time.monotonic() - last_send_timestamp
                logging.info("Elapsed since previous send: %.2f seconds", elapsed)
            last_send_timestamp = time.monotonic()
            if not dry_run and max_wait > 0:
                lower = max(0.0, min_wait)
                upper = max(lower, max_wait)
                wait_seconds = random.uniform(lower, upper)
                logging.info("Waiting %.2f seconds before next send", wait_seconds)
                time.sleep(wait_seconds)

        # Summary report
        logging.info("=" * 60)
        logging.info("Mailing Summary:")
        logging.info("  Total recipients: %d", total)
        logging.info("  Successfully sent: %d", success_count)
        logging.info("  Failed: %d", len(failed_recipients))

        if failed_recipients:
            logging.warning("=" * 60)
            logging.warning("Failed recipients:")
            for failed_email, error in failed_recipients:
                logging.warning("  - %s: %s", failed_email, error[:100])  # Truncate long errors
            logging.warning("=" * 60)


def _parse_recipients(
    xlsx_path: Path,
    *,
    sheet_name: Optional[str],
    mail_subject: str,
    sender_email: str = "",
) -> List[Recipient]:
    if not xlsx_path.exists():
        raise ConfigurationError(f"Spreadsheet file not found: {xlsx_path}")

    try:
        workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    except Exception as exc:  # pylint: disable=broad-except
        raise ConfigurationError(f"Unable to open spreadsheet: {xlsx_path}") from exc

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ConfigurationError(
                f"Worksheet '{sheet_name}' not found in {xlsx_path.name}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ConfigurationError("Spreadsheet contains no rows.")

    def _normalize(cell) -> str:
        if cell is None:
            return ""
        return str(cell).strip()

    # Build a header-name → column-index map from the first row.
    header_row = rows[0]
    col_map: Dict[str, int] = {
        _normalize(cell).lower(): idx
        for idx, cell in enumerate(header_row)
        if cell is not None
    }

    # Resolve column indices by header name.
    email_col = col_map.get("mail")
    name_col = col_map.get("name")
    login_id_col = col_map.get("login id")
    access_code_col = col_map.get("access code")
    subject_col = col_map.get("subject")

    if email_col is None:
        raise ConfigurationError(
            "Required column 'Mail' not found in spreadsheet header. "
            f"Found columns: {', '.join(_normalize(c) for c in header_row if c is not None)}"
        )

    recipients: List[Recipient] = []
    for idx, row in enumerate(rows[1:], start=2):
        def _get(col_idx: Optional[int]) -> str:
            if col_idx is None or col_idx >= len(row):
                return ""
            return _normalize(row[col_idx])

        recipient_email = _get(email_col)
        if not recipient_email:
            logging.warning("Row %d missing email; skipping.", idx)
            continue

        first_name = _get(name_col)
        login_id = _get(login_id_col)
        access_code = _get(access_code_col)

        if not login_id:
            pass
        if not access_code:
            pass

        subject = _get(subject_col) or _normalize(mail_subject)
        if not subject:
            raise ConfigurationError(
                f"Row {idx} has no subject. Add a 'Subject' column in the spreadsheet "
                "or provide --mail-subject / MAIL_SUBJECT as a fallback."
            )
        context = {
            "email": recipient_email,
            "first_name": first_name,
            "login_id": login_id,
            "access_code": access_code,
            "subject": subject,
            "sender_email": sender_email,
        }
        recipients.append(
            Recipient(
                email=recipient_email,
                first_name=first_name,
                subject=subject,
                context=context,
            )
        )

    if not recipients:
        raise ConfigurationError("No valid recipients found in spreadsheet.")
    return recipients


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Send personalized HTML emails with optional attachment via Microsoft 365."
    )
    default_save_to_sent = _env_flag("SAVE_TO_SENT_ITEMS", True)
    parser.add_argument(
        "--xlsx",
        dest="xlsx_path",
        default=os.getenv("XLSX_PATH", "mailing/recipients.xlsx"),
        help="Path to the recipient XLSX file (default: %(default)s).",
    )
    parser.add_argument(
        "--sheet-name",
        dest="sheet_name",
        default=os.getenv("RECIPIENT_SHEET_NAME"),
        help="Name of the worksheet to read (default: workbook's active sheet).",
    )
    parser.add_argument(
        "--template",
        dest="template_path",
        default=os.getenv("HTML_TEMPLATE_PATH", "mailing/email_template.html"),
        help="Path to the HTML template file (default: %(default)s).",
    )
    parser.add_argument(
        "--attachment",
        default=os.getenv("ATTACHMENT_PATH"),
        help="Path to the file attachment (optional).",
    )
    parser.add_argument(
        "--cc-email",
        dest="cc_email",
        default=os.getenv("CC_EMAIL"),
        help="Comma-separated email addresses to CC for every outgoing message (optional).",
    )
    parser.add_argument(
        "--mail-subject",
        default=os.getenv("MAIL_SUBJECT"),
        help="Fallback subject for rows that have no 'Subject' cell in the spreadsheet (default: MAIL_SUBJECT).",
    )
    parser.add_argument(
        "--log-level",
        default=os.getenv("LOG_LEVEL", "INFO"),
        help="Logging level (default: %(default)s).",
    )
    parser.add_argument(
        "--min-wait",
        type=float,
        default=float(os.getenv("MIN_WAIT_SECONDS", "5")),
        help="Minimum seconds to wait between messages (default: %(default)s).",
    )
    parser.add_argument(
        "--max-wait",
        type=float,
        default=float(os.getenv("MAX_WAIT_SECONDS", "15")),
        help="Maximum seconds to wait between messages (default: %(default)s).",
    )
    parser.add_argument(
        "--max-retries",
        type=int,
        default=int(os.getenv("MAX_RETRIES", "5")),
        help="Maximum retry attempts on transient errors (default: %(default)s).",
    )
    parser.add_argument(
        "--continue-on-error",
        dest="continue_on_error",
        action="store_true",
        default=_env_flag("CONTINUE_ON_ERROR", False),
        help="Continue sending to other recipients even if one fails (default: %(default)s).",
    )
    save_group = parser.add_mutually_exclusive_group()
    save_group.add_argument(
        "--save-to-sent-items",
        dest="save_to_sent_items",
        action="store_true",
        default=default_save_to_sent,
        help="Save outgoing messages to the Sent Items folder (default: enabled).",
    )
    save_group.add_argument(
        "--no-save-to-sent-items",
        dest="save_to_sent_items",
        action="store_false",
        help="Do not store outgoing messages in the Sent Items folder.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Render emails without calling the Graph API.",
    )
    return parser


def main() -> None:
    load_dotenv()
    parser = build_parser()
    args = parser.parse_args()

    logging.basicConfig(
        level=args.log_level.upper(),
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    try:
        tenant_id = _read_env("TENANT_ID")
        client_id = _read_env("CLIENT_ID")
        client_secret = _read_env("CLIENT_SECRET")
        sender_address = _read_env("SENDER_EMAIL")

        template_path = Path(args.template_path)
        html_template = _load_html_template(template_path)
        html_template, inline_attachments = _prepare_inline_images(
            html_template, template_path.parent
        )


        recipients = _parse_recipients(
            Path(args.xlsx_path),
            sheet_name=args.sheet_name,
            mail_subject=args.mail_subject,
            sender_email=sender_address,
        )

        attachment = None
        if args.attachment:
            attachment = _build_attachment(Path(args.attachment))

        mailer = GraphMailer(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret,
            sender=sender_address,
        )
        mailer.send(
            recipients,
            html_template,
            inline_attachments=inline_attachments,
            attachment=attachment,
            min_wait=args.min_wait,
            max_wait=args.max_wait,
            max_retries=args.max_retries,
            dry_run=args.dry_run,
            save_to_sent_items=args.save_to_sent_items,
            cc_email=(args.cc_email.strip() if args.cc_email else None),
            continue_on_error=args.continue_on_error,
        )
    except ConfigurationError as exc:
        logging.error("%s", exc)
    except Exception as exc:  # pylint: disable=broad-except
        logging.exception("Unexpected error: %s", exc)


if __name__ == "__main__":
    main()
