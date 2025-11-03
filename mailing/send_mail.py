"""
Utility for sending personalized HTML emails with attachments via Microsoft 365.

The script reads recipient data from an XLSX spreadsheet, renders an HTML template using
row values, and delivers the messages through the Microsoft Graph API.
"""

from __future__ import annotations

import argparse
import base64
import json
import logging
import mimetypes
import os
import random
import re
import time
import uuid
from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import requests
from openpyxl import load_workbook
from premailer import transform
from dotenv import load_dotenv
from msal import ConfidentialClientApplication

GRAPH_SCOPE = ["https://graph.microsoft.com/.default"]
GRAPH_ENDPOINT = "https://graph.microsoft.com/v1.0"


class ConfigurationError(RuntimeError):
    """Raised when required configuration is missing."""


def _read_env(name: str, *, required: bool = True, default: Optional[str] = None) -> str:
    value = os.getenv(name, default)
    if required and not value:
        raise ConfigurationError(
            f"Missing required configuration for {name}. "
            "Set it in your environment or .env file."
        )
    return value or ""


def _env_flag(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _load_html_template(path: Path) -> str:
    try:
        with open(path, 'r') as f:
            html = f.read()

        inline_html = transform(html)

        with open(path, 'w') as f:
            f.write(inline_html)
        
        return path.read_text(encoding="utf-8")
    except FileNotFoundError as exc:
        raise ConfigurationError(f"HTML template not found: {path}") from exc


def _guess_mime_type(path: Path) -> str:
    mime, _ = mimetypes.guess_type(path.name)
    return mime or "application/octet-stream"


def _build_attachment(path: Path) -> Dict[str, str]:
    if not path.exists():
        raise ConfigurationError(f"Attachment file does not exist: {path}")
    file_bytes = path.read_bytes()
    content_bytes = base64.b64encode(file_bytes).decode("ascii")
    return {
        "@odata.type": "#microsoft.graph.fileAttachment",
        "name": path.name,
        "contentType": _guess_mime_type(path),
        "contentBytes": content_bytes,
    }


def _prepare_inline_images(html: str, asset_root: Path) -> Tuple[str, List[Dict[str, str]]]:
    """
    Identify local <img> references and convert them to inline attachments that
    Microsoft Graph can embed via CID references.
    """
    matches = set(re.findall(r'<img[^>]+src=["\']([^"\']+)["\']', html, flags=re.IGNORECASE))
    if not matches:
        return html, []

    replacements: Dict[str, str] = {}
    attachments: List[Dict[str, str]] = []

    for src in matches:
        if src.startswith(("cid:", "http://", "https://", "data:")):
            continue
        image_path = Path(src)
        if not image_path.is_absolute():
            image_path = asset_root / src
        if not image_path.exists():
            logging.warning("Referenced inline image not found: %s", image_path)
            continue
        content_id = f"{Path(src).stem}-{uuid.uuid4().hex}@inline"
        file_bytes = image_path.read_bytes()
        attachments.append(
            {
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": image_path.name,
                "contentType": _guess_mime_type(image_path),
                "contentBytes": base64.b64encode(file_bytes).decode("ascii"),
                "contentId": content_id,
                "isInline": True,
            }
        )
        replacements[src] = content_id

    if not replacements:
        return html, attachments

    def _replace(match: re.Match) -> str:
        prefix, quote, value, _ = match.groups()
        content_id = replacements.get(value)
        if not content_id:
            return match.group(0)
        return f"{prefix}{quote}cid:{content_id}{quote}"

    updated_html = re.sub(
        r'(<img[^>]*src=)(["\'])([^"\']+)(\2)',
        _replace,
        html,
        flags=re.IGNORECASE,
    )
    return updated_html, attachments


@dataclass
class Recipient:
    email: str
    first_name: str
    subject: str
    context: Dict[str, str]


class GraphMailer:
    def __init__(self, tenant_id: str, client_id: str, client_secret: str, sender: str) -> None:
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        self._sender = sender
        self._client = ConfidentialClientApplication(
            client_id=client_id,
            client_credential=client_secret,
            authority=authority,
        )

    def _get_token(self) -> str:
        token = self._client.acquire_token_silent(GRAPH_SCOPE, account=None)
        if not token:
            token = self._client.acquire_token_for_client(scopes=GRAPH_SCOPE)
        if "access_token" not in token:
            raise RuntimeError(f"Unable to acquire access token: {json.dumps(token, indent=2)}")
        return token["access_token"]

    def send(
        self,
        recipients: Iterable[Recipient],
        html_template: str,
        *,
        inline_attachments: Optional[List[Dict[str, str]]] = None,
        attachment: Optional[Dict[str, str]] = None,
        min_wait: float = 5.0,
        max_wait: float = 15.0,
        dry_run: bool = False,
        save_to_sent_items: bool = True,
    ) -> None:
        token = None if dry_run else self._get_token()
        total = 0
        inline_attachments = inline_attachments or []
        last_send_timestamp: Optional[float] = None
        for recipient in recipients:
            total += 1
            rendered_html = html_template.format(**recipient.context)
            if dry_run:
                logging.info(
                    "[DRY-RUN] Would send: \n Subject: '%s' - Email: %s - Name: [%s]",
                    recipient.subject,
                    recipient.email,
                    recipient.first_name,
                )
                continue
            payload = {
                "message": {
                    "subject": recipient.subject,
                    "body": {"contentType": "HTML", "content": rendered_html},
                    "toRecipients": [{"emailAddress": {"address": recipient.email}}],
                },
                "saveToSentItems": save_to_sent_items,
            }
            message_attachments: List[Dict[str, str]] = []
            if inline_attachments:
                message_attachments.extend(deepcopy(inline_attachments))
            if attachment:
                message_attachments.append(deepcopy(attachment))
            if message_attachments:
                payload["message"]["attachments"] = message_attachments
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            }
            response = requests.post(
                f"{GRAPH_ENDPOINT}/users/{self._sender}/sendMail",
                headers=headers,
                json=payload,
                timeout=30,
            )
            if response.status_code >= 300:
                raise RuntimeError(
                    f"Failed to send mail to {recipient.email}: "
                    f"{response.status_code} {response.text}"
                )
            logging.info("Sent mail to %s", recipient.email)
            if last_send_timestamp is not None:
                elapsed = time.monotonic() - last_send_timestamp
                logging.info("Elapsed since previous send: %.2f seconds", elapsed)
            last_send_timestamp = time.monotonic()
            if not dry_run and max_wait > 0:
                lower = max(0.0, min_wait)
                upper = max(lower, max_wait)
                wait_seconds = random.uniform(lower, upper)
                logging.info("Waiting %.2f seconds before next send", wait_seconds)
                time.sleep(wait_seconds)
        logging.info("Processed %d recipient(s).", total)


def _parse_recipients(
    xlsx_path: Path,
    *,
    sheet_name: Optional[str],
    default_subject: Optional[str],
) -> List[Recipient]:
    if not xlsx_path.exists():
        raise ConfigurationError(f"Spreadsheet file not found: {xlsx_path}")

    try:
        workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    except Exception as exc:  # pylint: disable=broad-except
        raise ConfigurationError(f"Unable to open spreadsheet: {xlsx_path}") from exc

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ConfigurationError(
                f"Worksheet '{sheet_name}' not found in {xlsx_path.name}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ConfigurationError("Spreadsheet contains no rows.")

    expected_order = ("email", "first_name", "sender_name", "subject")
    header_row = rows[0]

    def _normalize(cell: Optional[str]) -> str:
        if cell is None:
            return ""
        return str(cell).strip()

    header_matches = all(
        _normalize(header_row[idx]).lower() == expected_order[idx]
        for idx in range(min(len(expected_order), len(header_row)))
    )
    data_rows = rows[1:] if header_matches else rows

    recipients: List[Recipient] = []
    starting_index = 2 if header_matches else 1
    for idx, row in enumerate(data_rows, start=starting_index):
        # Pad the row up to 4 entries to guard against shorter rows.
        padded = list(row) + [None] * max(0, 4 - len(row))
        email = _normalize(padded[0])
        if not email:
            logging.warning("Row %d missing email; skipping.", idx)
            continue
        first_name = _normalize(padded[1])
        sender_name = _normalize(padded[2])
        subject = _normalize(padded[3]) or _normalize(default_subject)
        if not subject:
            raise ConfigurationError(
                f"Row {idx} missing subject and no default subject provided."
            )
        context = {
            "email": email,
            "first_name": first_name,
            "sender_name": sender_name,
            "subject": subject,
        }
        recipients.append(
            Recipient(
                email=email,
                first_name=first_name,
                subject=subject,
                context=context,
            )
        )
    if not recipients:
        raise ConfigurationError("No valid recipients found in spreadsheet.")
    return recipients


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Send personalized HTML emails with optional attachment via Microsoft 365."
    )
    default_save_to_sent = _env_flag("SAVE_TO_SENT_ITEMS", True)
    parser.add_argument(
        "--xlsx",
        dest="xlsx_path",
        default=os.getenv("XLSX_PATH", "mailing/recipients.xlsx"),
        help="Path to the recipient XLSX file (default: %(default)s).",
    )
    parser.add_argument(
        "--sheet-name",
        dest="sheet_name",
        default=os.getenv("RECIPIENT_SHEET_NAME"),
        help="Name of the worksheet to read (default: workbook's active sheet).",
    )
    parser.add_argument(
        "--template",
        dest="template_path",
        default=os.getenv("HTML_TEMPLATE_PATH", "mailing/email_template.html"),
        help="Path to the HTML template file (default: %(default)s).",
    )
    parser.add_argument(
        "--attachment",
        default=os.getenv("ATTACHMENT_PATH"),
        help="Path to the file attachment (optional).",
    )
    parser.add_argument(
        "--default-subject",
        default=os.getenv("DEFAULT_SUBJECT"),
        help="Subject to use when the subject column is absent or empty.",
    )
    parser.add_argument(
        "--log-level",
        default=os.getenv("LOG_LEVEL", "INFO"),
        help="Logging level (default: %(default)s).",
    )
    parser.add_argument(
        "--min-wait",
        type=float,
        default=float(os.getenv("MIN_WAIT_SECONDS", "5")),
        help="Minimum seconds to wait between messages (default: %(default)s).",
    )
    parser.add_argument(
        "--max-wait",
        type=float,
        default=float(os.getenv("MAX_WAIT_SECONDS", "15")),
        help="Maximum seconds to wait between messages (default: %(default)s).",
    )
    save_group = parser.add_mutually_exclusive_group()
    save_group.add_argument(
        "--save-to-sent-items",
        dest="save_to_sent_items",
        action="store_true",
        default=default_save_to_sent,
        help="Save outgoing messages to the Sent Items folder (default: enabled).",
    )
    save_group.add_argument(
        "--no-save-to-sent-items",
        dest="save_to_sent_items",
        action="store_false",
        help="Do not store outgoing messages in the Sent Items folder.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Render emails without calling the Graph API.",
    )
    return parser


def main() -> None:
    load_dotenv()
    parser = build_parser()
    args = parser.parse_args()

    logging.basicConfig(
        level=args.log_level.upper(),
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    try:
        tenant_id = _read_env("TENANT_ID")
        client_id = _read_env("CLIENT_ID")
        client_secret = _read_env("CLIENT_SECRET")
        sender_address = _read_env("SENDER_EMAIL")

        template_path = Path(args.template_path)
        html_template = _load_html_template(template_path)
        html_template, inline_attachments = _prepare_inline_images(
            html_template, template_path.parent
        )

        recipients = _parse_recipients(
            Path(args.xlsx_path),
            sheet_name=args.sheet_name,
            default_subject=args.default_subject,
        )

        attachment = None
        if args.attachment:
            attachment = _build_attachment(Path(args.attachment))

        mailer = GraphMailer(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret,
            sender=sender_address,
        )
        mailer.send(
            recipients,
            html_template,
            inline_attachments=inline_attachments,
            attachment=attachment,
            min_wait=args.min_wait,
            max_wait=args.max_wait,
            dry_run=args.dry_run,
            save_to_sent_items=args.save_to_sent_items,
        )
    except ConfigurationError as exc:
        logging.error("%s", exc)
    except Exception as exc:  # pylint: disable=broad-except
        logging.exception("Unexpected error: %s", exc)


if __name__ == "__main__":
    main()
