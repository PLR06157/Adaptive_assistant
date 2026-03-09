"""
Send personalized reading assessment results to each participant.

Reads Assessment_mails.xlsx (Name = col B, Mail = col E, Login ID = col F),
parses reading_assessment_individual_reports.md, converts each participant's
section to styled HTML, and delivers via Microsoft Graph API.

GraphMailer and all Graph/auth logic are reused from send_mail.py.

Usage:

  Dry run (no emails sent):
    python3 send_assessment.py --xlsx sets/2026_02_Amway_trng/Assessment_mails_test.xlsx --report sets/2026_02_Amway_trng/reading_assessment_individual_reports.md --template sets/2026_02_Amway_trng/assessment_email.html --mail-subject "AI Training Homework - your results" --dry-run

  Live send:
    python3 send_assessment.py --xlsx sets/2026_02_Amway_trng/Assessment_mails.xlsx --report sets/2026_02_Amway_trng/reading_assessment_individual_reports.md --template sets/2026_02_Amway_trng/assessment_email.html --mail-subject "AI Training Homework - your results"

Configuration:
    Reads the same .env file as send_mail.py (must be in mailing/ directory).
    Required: TENANT_ID, CLIENT_ID, CLIENT_SECRET, SENDER_EMAIL
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook

# ── Import shared components from send_mail.py (same directory) ───────────────
sys.path.insert(0, str(Path(__file__).parent))
from send_mail import (  # noqa: E402
    ConfigurationError,
    GraphMailer,
    Recipient,
    _env_flag,
    _prepare_inline_images,
    _read_env,
)

# ── Colour palette (mirrors assessment_email.html) ────────────────────────────
_GOLD = "#B8975A"
_DARK = "#2C2C2C"
_BG = "#F8F6F3"


# ═════════════════════════════════════════════════════════════════════════════
# MD Parsing
# ═════════════════════════════════════════════════════════════════════════════

def parse_participant_sections(md_path: Path) -> Dict[str, str]:
    """
    Return {USER_CODE: section_body} for every participant in the MD report.
    Splits on '## Participant: USER-XXXXX' headings.
    """
    text = md_path.read_text(encoding="utf-8")
    parts = re.split(r"^## Participant:\s*(USER-\w+)", text, flags=re.MULTILINE)
    # parts[0] = preamble; then alternating: user_code, body_text
    sections: Dict[str, str] = {}
    for i in range(1, len(parts), 2):
        user_code = parts[i].strip()
        body = parts[i + 1].strip() if i + 1 < len(parts) else ""
        sections[user_code] = body
    return sections


# ═════════════════════════════════════════════════════════════════════════════
# MD → HTML Conversion
# ═════════════════════════════════════════════════════════════════════════════

def _render_score(score: int, score_max: int) -> str:
    return f"""
<table role="presentation" width="100%" border="0" cellpadding="0" cellspacing="0"
       style="background-color: {_BG}; border: 2px solid {_GOLD}; border-radius: 6px; margin-bottom: 20px;">
  <tr>
    <td style="padding: 18px; text-align: center;">
      <p style="margin: 0; font-size: 11px; color: #888; text-transform: uppercase; letter-spacing: 1px; font-family: 'Open Sans', Arial, sans-serif;">
        Overall Score
      </p>
      <p style="margin: 6px 0 0 0; font-size: 42px; font-weight: 700; color: {_GOLD}; font-family: Copperplate, Georgia, serif; line-height: 1;">
        {score}
        <span style="font-size: 20px; color: #aaa; font-family: 'Open Sans', Arial, sans-serif;">/ {score_max}</span>
      </p>
    </td>
  </tr>
</table>"""


def _render_placement(placement: str) -> str:
    colors: Dict[str, Tuple[str, str, str]] = {
        "Strong":   ("#1B5E20", "#E8F5E9", "#4CAF50"),
        "Adequate": ("#7B5000", "#FFF8E1", "#FFC107"),
        "Weak":     ("#B71C1C", "#FFEBEE", "#E53935"),
    }
    text_c, bg_c, border_c = colors.get(placement, (_DARK, _BG, _GOLD))
    return f"""
<table role="presentation" border="0" cellpadding="0" cellspacing="0" align="center"
       style="margin: 0 auto 20px auto;">
  <tr>
    <td style="background-color: {bg_c}; border: 1px solid {border_c}; border-radius: 4px; padding: 8px 24px;">
      <span style="font-size: 12px; color: {text_c}; font-weight: 700; text-transform: uppercase; letter-spacing: 1px; font-family: 'Open Sans', Arial, sans-serif;">
        Placement Signal: {placement}
      </span>
    </td>
  </tr>
</table>"""


def _render_notes(notes: str) -> str:
    return (
        f'<p style="font-size: 14px; color: {_DARK}; margin: 0 0 20px 0;'
        f" font-family: 'Open Sans', Arial, sans-serif;\">{notes}</p>"
    )


def _render_breakdown(items: List[Tuple[str, str]]) -> str:
    if not items:
        return ""
    rows = ""
    for criterion_score, desc in items:
        rows += f"""
    <tr>
      <td style="padding: 9px 12px; border-bottom: 1px solid #E8E0D5; font-size: 13px; color: {_DARK}; font-family: 'Open Sans', Arial, sans-serif;">
        {criterion_score}
      </td>
      <td style="padding: 9px 12px; border-bottom: 1px solid #E8E0D5; font-size: 13px; color: #555; font-family: 'Open Sans', Arial, sans-serif;">
        {desc}
      </td>
    </tr>"""
    return f"""
<p style="font-size: 14px; font-weight: 700; color: {_DARK}; margin: 0 0 10px 0; font-family: 'Open Sans', Arial, sans-serif;">
  Score Breakdown
</p>
<table role="presentation" width="100%" border="0" cellpadding="0" cellspacing="0"
       style="border: 1px solid #E8E0D5; border-radius: 6px; border-collapse: collapse; margin-bottom: 20px;">
  <tr style="background-color: {_BG};">
    <th style="padding: 8px 12px; font-size: 12px; text-align: left; color: #777; font-weight: 600; border-bottom: 2px solid #E8E0D5; font-family: 'Open Sans', Arial, sans-serif; width: 45%;">
      Criterion
    </th>
    <th style="padding: 8px 12px; font-size: 12px; text-align: left; color: #777; font-weight: 600; border-bottom: 2px solid #E8E0D5; font-family: 'Open Sans', Arial, sans-serif;">
      Evaluator Note
    </th>
  </tr>
  {rows}
</table>"""


def _render_feedback(
    keep_doing: List[str],
    improve_next: List[str],
    suggested_practice: str,
) -> str:
    if not (keep_doing or improve_next or suggested_practice):
        return ""

    content = ""
    if keep_doing:
        items_html = "".join(
            f'<li style="margin-bottom: 5px;">{item}</li>' for item in keep_doing
        )
        content += (
            f'<p style="font-size: 13px; font-weight: 700; color: #2E7D32; margin: 0 0 6px 0;'
            f" font-family: 'Open Sans', Arial, sans-serif;\">Keep doing</p>"
            f'<ul style="margin: 0 0 14px 0; padding-left: 18px; font-size: 13px; color: {_DARK};'
            f" font-family: 'Open Sans', Arial, sans-serif;\">{items_html}</ul>"
        )
    if improve_next:
        items_html = "".join(
            f'<li style="margin-bottom: 5px;">{item}</li>' for item in improve_next
        )
        content += (
            f'<p style="font-size: 13px; font-weight: 700; color: #E65100; margin: 0 0 6px 0;'
            f" font-family: 'Open Sans', Arial, sans-serif;\">Improve next</p>"
            f'<ul style="margin: 0 0 14px 0; padding-left: 18px; font-size: 13px; color: {_DARK};'
            f" font-family: 'Open Sans', Arial, sans-serif;\">{items_html}</ul>"
        )
    if suggested_practice:
        content += f"""
<table role="presentation" width="100%" border="0" cellpadding="0" cellspacing="0"
       style="background-color: #FFF8E1; border: 1px solid #FFC107; border-radius: 6px; margin-top: 4px;">
  <tr>
    <td style="padding: 12px 16px;">
      <p style="margin: 0 0 4px 0; font-size: 11px; font-weight: 700; color: #7B5000; text-transform: uppercase; letter-spacing: 1px; font-family: 'Open Sans', Arial, sans-serif;">
        Suggested practice for next week
      </p>
      <p style="margin: 0; font-size: 13px; color: {_DARK}; font-family: 'Open Sans', Arial, sans-serif;">
        {suggested_practice}
      </p>
    </td>
  </tr>
</table>"""

    return f"""
<table role="presentation" width="100%" border="0" cellpadding="0" cellspacing="0"
       style="background-color: #FAFAFA; border: 1px solid #E8E0D5; border-radius: 6px; margin-bottom: 20px;">
  <tr>
    <td style="padding: 16px 18px;">
      <p style="font-size: 14px; font-weight: 700; color: {_DARK}; margin: 0 0 14px 0; font-family: 'Open Sans', Arial, sans-serif;">
        Individual Feedback
      </p>
      {content}
    </td>
  </tr>
</table>"""


def section_to_html(section_text: str) -> str:
    """Convert one participant's MD section body to inline-styled HTML."""

    # ── Score ────────────────────────────────────────────────────────────────
    score_m = re.search(r"\*\*Score:\s*(\d+)\s*/\s*(\d+)\*\*", section_text)
    score = int(score_m.group(1)) if score_m else 0
    score_max = int(score_m.group(2)) if score_m else 10

    # ── Breakdown ────────────────────────────────────────────────────────────
    breakdown_items: List[Tuple[str, str]] = []
    bd_m = re.search(r"\*\*Breakdown:\*\*\n((?:- .+\n?)+)", section_text)
    if bd_m:
        for line in bd_m.group(1).strip().splitlines():
            line = line.strip().lstrip("- ").strip()
            if " \u2014 " in line:  # em-dash separator used in the MD
                crit, desc = line.split(" \u2014 ", 1)
                breakdown_items.append((crit.strip(), desc.strip()))
            elif line:
                breakdown_items.append((line, ""))

    # ── Placement Signal ─────────────────────────────────────────────────────
    pl_m = re.search(r"\*\*Placement Signal:\*\*\s*(\w+)", section_text)
    placement = pl_m.group(1).strip() if pl_m else ""

    # ── Evaluator Notes ──────────────────────────────────────────────────────
    notes_m = re.search(
        r"\*\*Brief Evaluator Notes:\*\*\s*(.+?)(?=\n\*\*|\Z)", section_text, re.DOTALL
    )
    notes = notes_m.group(1).strip() if notes_m else ""

    # ── Individual Feedback ──────────────────────────────────────────────────
    keep_doing: List[str] = []
    improve_next: List[str] = []
    suggested_practice = ""

    fb_m = re.search(
        r"\*\*Individual feedback \(actionable\):\*\*\n(.*?)(?=\n\*\*|\Z)",
        section_text,
        re.DOTALL,
    )
    if fb_m:
        fb = fb_m.group(1)
        kd_m = re.search(r"- Keep doing:\n((?:  - .+\n?)+)", fb)
        if kd_m:
            for line in kd_m.group(1).strip().splitlines():
                keep_doing.append(line.strip().lstrip("- ").strip())
        in_m = re.search(r"- Improve next:\n((?:  - .+\n?)+)", fb)
        if in_m:
            for line in in_m.group(1).strip().splitlines():
                improve_next.append(line.strip().lstrip("- ").strip())
        sp_m = re.search(
            r"- Suggested practice for next week:\s*(.+?)(?=\n-|\Z)", fb, re.DOTALL
        )
        if sp_m:
            suggested_practice = sp_m.group(1).strip()

    # ── Assemble ─────────────────────────────────────────────────────────────
    parts = [_render_score(score, score_max)]
    if placement:
        parts.append(_render_placement(placement))
    if notes:
        parts.append(_render_notes(notes))
    parts.append(_render_breakdown(breakdown_items))
    parts.append(_render_feedback(keep_doing, improve_next, suggested_practice))

    return "\n".join(p for p in parts if p)


# ═════════════════════════════════════════════════════════════════════════════
# XLSX Loading
# ═════════════════════════════════════════════════════════════════════════════

# Fixed column positions (0-based): B=1, E=4, F=5
_COL_NAME = 1
_COL_MAIL = 4
_COL_LOGIN = 5


def _load_recipients(
    xlsx_path: Path,
    mail_subject: str,
    sender_email: str,
) -> List[Recipient]:
    """
    Read Assessment_mails.xlsx.
    Row 1 is treated as the header and skipped.
    Columns: Name=B (idx 1), Mail=E (idx 4), Login ID / USER code=F (idx 5).
    """
    if not xlsx_path.exists():
        raise ConfigurationError(f"Spreadsheet not found: {xlsx_path}")
    try:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ConfigurationError(f"Cannot open spreadsheet: {xlsx_path}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ConfigurationError("Spreadsheet has no data rows (expected header + data).")

    header = rows[0]
    logging.debug("XLSX header row: %s", header)

    def _cell(row: tuple, idx: int) -> str:
        if idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    recipients: List[Recipient] = []
    for row_num, row in enumerate(rows[1:], start=2):
        email = _cell(row, _COL_MAIL)
        if not email:
            logging.warning("Row %d: no email address — skipping.", row_num)
            continue
        name = _cell(row, _COL_NAME)
        user_code = _cell(row, _COL_LOGIN)
        if not user_code:
            logging.warning("Row %d (%s): no Login ID / USER code — skipping.", row_num, email)
            continue

        context = {
            "email": email,
            "first_name": name,
            "user_code": user_code,
            "subject": mail_subject,
            "sender_email": sender_email,
            "assessment_html": "",  # populated after MD matching
        }
        recipients.append(
            Recipient(email=email, first_name=name, subject=mail_subject, context=context)
        )

    if not recipients:
        raise ConfigurationError("No valid recipients found in spreadsheet.")
    return recipients


# ═════════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════════

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Send individual reading assessment results to each participant."
    )
    p.add_argument(
        "--xlsx",
        default="mailing/sets/2026_02_Amway_trng/Assessment_mails.xlsx",
        help="Path to Assessment_mails.xlsx (default: %(default)s).",
    )
    p.add_argument(
        "--report",
        default="mailing/sets/2026_02_Amway_trng/reading_assessment_individual_reports.md",
        help="Path to the consolidated MD assessment report (default: %(default)s).",
    )
    p.add_argument(
        "--template",
        default="mailing/sets/2026_02_Amway_trng/assessment_email.html",
        help="Path to the HTML email template (default: %(default)s).",
    )
    p.add_argument(
        "--mail-subject",
        default=os.getenv("MAIL_SUBJECT", ""),
        help="Email subject line (also settable via MAIL_SUBJECT env var).",
    )
    p.add_argument(
        "--cc-email",
        dest="cc_email",
        default=os.getenv("CC_EMAIL"),
        help="Comma-separated email addresses to CC on every outgoing message (optional).",
    )
    p.add_argument(
        "--min-wait",
        type=float,
        default=float(os.getenv("MIN_WAIT_SECONDS", "5")),
        help="Minimum seconds between sends (default: %(default)s).",
    )
    p.add_argument(
        "--max-wait",
        type=float,
        default=float(os.getenv("MAX_WAIT_SECONDS", "15")),
        help="Maximum seconds between sends (default: %(default)s).",
    )
    p.add_argument(
        "--max-retries",
        type=int,
        default=int(os.getenv("MAX_RETRIES", "5")),
        help="Max retry attempts on transient errors (default: %(default)s).",
    )
    p.add_argument(
        "--log-level",
        default=os.getenv("LOG_LEVEL", "INFO"),
        help="Logging level (default: %(default)s).",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be sent without calling the Graph API.",
    )
    p.add_argument(
        "--continue-on-error",
        dest="continue_on_error",
        action="store_true",
        default=_env_flag("CONTINUE_ON_ERROR", False),
        help="Continue sending to remaining recipients if one fails.",
    )
    default_save = _env_flag("SAVE_TO_SENT_ITEMS", True)
    save_group = p.add_mutually_exclusive_group()
    save_group.add_argument(
        "--save-to-sent-items",
        dest="save_to_sent_items",
        action="store_true",
        default=default_save,
    )
    save_group.add_argument(
        "--no-save-to-sent-items",
        dest="save_to_sent_items",
        action="store_false",
    )
    return p


def main() -> None:
    load_dotenv()
    parser = build_parser()
    args = parser.parse_args()

    logging.basicConfig(
        level=args.log_level.upper(),
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    try:
        if not args.mail_subject:
            raise ConfigurationError(
                "--mail-subject is required (or set MAIL_SUBJECT in .env)."
            )

        tenant_id = _read_env("TENANT_ID")
        client_id = _read_env("CLIENT_ID")
        client_secret = _read_env("CLIENT_SECRET")
        sender_address = _read_env("SENDER_EMAIL")

        # ── Load MD report ────────────────────────────────────────────────────
        report_path = Path(args.report)
        if not report_path.exists():
            raise ConfigurationError(f"Report file not found: {report_path}")
        sections = parse_participant_sections(report_path)
        logging.info("Loaded %d participant sections from %s", len(sections), report_path.name)

        # ── Load XLSX recipients ──────────────────────────────────────────────
        recipients = _load_recipients(Path(args.xlsx), args.mail_subject, sender_address)
        logging.info("Loaded %d recipients from XLSX", len(recipients))

        # ── Load HTML template ────────────────────────────────────────────────
        template_path = Path(args.template)
        if not template_path.exists():
            raise ConfigurationError(f"Template not found: {template_path}")
        html_template = template_path.read_text(encoding="utf-8")

        # Convert local images to CID inline attachments
        html_template, inline_attachments = _prepare_inline_images(
            html_template, template_path.parent
        )

        # ── Match recipients to assessment sections ───────────────────────────
        matched, skipped = 0, []
        for r in recipients:
            user_code = r.context["user_code"]
            if user_code not in sections:
                logging.warning(
                    "No section found for %s (%s) — will skip.", user_code, r.email
                )
                skipped.append(r.email)
                continue
            r.context["assessment_html"] = section_to_html(sections[user_code])
            matched += 1

        # Drop recipients that had no matching section
        recipients = [r for r in recipients if r.context["assessment_html"]]
        logging.info(
            "%d recipients matched; %d skipped (no section in MD).", matched, len(skipped)
        )
        if skipped:
            logging.warning("Skipped emails: %s", ", ".join(skipped))
        if not recipients:
            raise ConfigurationError("No recipients remaining after matching. Aborting.")

        # ── Send ──────────────────────────────────────────────────────────────
        mailer = GraphMailer(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret,
            sender=sender_address,
        )
        mailer.send(
            recipients,
            html_template,
            inline_attachments=inline_attachments,
            min_wait=args.min_wait,
            max_wait=args.max_wait,
            max_retries=args.max_retries,
            dry_run=args.dry_run,
            save_to_sent_items=args.save_to_sent_items,
            cc_email=(args.cc_email.strip() if args.cc_email else None),
            continue_on_error=args.continue_on_error,
        )

    except ConfigurationError as exc:
        logging.error("%s", exc)
        sys.exit(1)
    except Exception as exc:  # pylint: disable=broad-except
        logging.exception("Unexpected error: %s", exc)
        sys.exit(1)


if __name__ == "__main__":
    main()
